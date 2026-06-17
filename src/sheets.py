import logging
import os

import pandas as pd

from config import (
    CSV_SEPARATOR,
    CSV_DECIMAL,
    REPORTS_DIR,
)

logger = logging.getLogger(__name__)
os.makedirs(REPORTS_DIR, exist_ok=True)

def consolidate_reports(
    all_descriptive: list[pd.DataFrame],
    all_hypothesis:  list[pd.DataFrame],
) -> dict[str, str]:
    """
    Concatena os DataFrames de todos os parceiros e exporta relatórios
    consolidados.

    """
    output_paths: dict[str, str] = {}

    valid_desc = [df for df in all_descriptive if df is not None and not df.empty]
    if valid_desc:
        desc_concat = pd.concat(valid_desc, ignore_index=True)
        path_desc = os.path.join(REPORTS_DIR, "consolidado_descritivo.csv")
        desc_concat.to_csv(
            path_desc,
            sep=CSV_SEPARATOR,
            decimal=CSV_DECIMAL,
            index=False,
            encoding="utf-8-sig",
        )
        output_paths["csv_descritivo"] = path_desc
        logger.info("CSV consolidado descritivo: %s", path_desc)
    else:
        logger.warning("Nenhum dado descritivo para consolidar.")

    valid_hyp = [df for df in all_hypothesis if df is not None and not df.empty]
    if valid_hyp:
        hyp_concat = pd.concat(valid_hyp, ignore_index=True)
        path_hyp = os.path.join(REPORTS_DIR, "consolidado_hipoteses.csv")
        hyp_concat.to_csv(
            path_hyp,
            sep=CSV_SEPARATOR,
            decimal=CSV_DECIMAL,
            index=False,
            encoding="utf-8-sig",
        )
        output_paths["csv_hipoteses"] = path_hyp
        logger.info("CSV consolidado hipóteses: %s", path_hyp)
    else:
        logger.warning("Nenhum dado de hipóteses para consolidar.")

    path_xlsx = os.path.join(REPORTS_DIR, "consolidado_resumo.xlsx")
    try:
        with pd.ExcelWriter(path_xlsx, engine="openpyxl") as writer:
            if valid_desc:
                desc_concat.to_excel(
                    writer,
                    sheet_name="Descritivo",
                    index=False,
                )
                _format_excel_sheet(writer, "Descritivo", desc_concat)

            if valid_hyp:
                hyp_concat.to_excel(
                    writer,
                    sheet_name="Testes_Hipotese",
                    index=False,
                )
                _format_excel_sheet(writer, "Testes_Hipotese", hyp_concat)

            if valid_hyp:
                pivot_df = _build_kpi_pivot(hyp_concat)
                if pivot_df is not None and not pivot_df.empty:
                    pivot_df.to_excel(writer, sheet_name="KPI_Pivot")
                    _format_excel_sheet(writer, "KPI_Pivot", pivot_df, is_pivot=True)

        output_paths["xlsx"] = path_xlsx
        logger.info("Excel consolidado: %s", path_xlsx)
    except ImportError:
        logger.warning("openpyxl não instalado — Excel não gerado. Execute: pip install openpyxl")
    except Exception as e:
        logger.error("Erro ao gerar Excel: %s", e)

    return output_paths

def _build_kpi_pivot(hyp_df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Constrói uma tabela pivot das métricas-chave:
    Parceiro × Grupo Variante × Uplift relativo por métrica.
    """
    try:
        pivot = hyp_df.pivot_table(
            index=["parceiro", "grupo_variante"],
            columns="metrica",
            values="uplift_relativo_pct",
            aggfunc="first",
        )
        pivot.columns.name = None
        pivot.reset_index(inplace=True)
        return pivot
    except Exception as e:
        logger.warning("Não foi possível criar pivot KPI: %s", e)
        return None

def _format_excel_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    is_pivot: bool = False,
) -> None:
    """
    Aplica formatação básica à planilha Excel:
    ajuste de largura das colunas e destaque do cabeçalho.
    Requer openpyxl.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(color="E5E7EB", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value), default=8
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    except Exception as e:
        logger.debug("Formatação Excel falhou: %s", e)